#!/usr/bin/python

#import pylab as pl
import numpy as np
from scipy.optimize import fsolve
from GreeceModelRBFOC import GreeceModelRBFOC
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
import sys, os
from Tkinter import *
import timeit
import tkMessageBox
import matplotlib
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
from openpyxl.compat import range
from openpyxl.utils import get_column_letter



wb = load_workbook(filename = 'MoU.xlsx')
count = 1
#Buttons government
def click_tau_k():
    try:
        tau_k = float(entrada_tau_k.get())
        etiqueta_tau_k.config(text = "t_k = "+ "%.3f" %tau_k)
    except ValueError:
        etiqueta_tau_k.config(text = "InvaLid value")
        
def click_tau_l():
    try:
        tau_l = float(entrada_tau_l.get())
        etiqueta_tau_l.config(text = "t_l = "+ "%.3f" %tau_l)
    except ValueError:
        etiqueta_tau_l.config(text = "Invalid value")
        
def click_tau_c():
    try:
        tau_c = float(entrada_tau_c.get())
        etiqueta_tau_c.config(text = "t_c = "+ "%.3f" %tau_c)
    except ValueError:
        etiqueta_tau_c.config(text = "Invalid value")
        
def click_tau_ss():
    try:
        tau_ss = float(entrada_tau_ss.get())
        etiqueta_tau_ss.config(text = "t_ss = "+ "%.3f" %tau_ss)
    except ValueError:
        etiqueta_tau_ss.config(text = "Invalid value")
        
def click_tau_pi():
    try:
        tau_pi = float(entrada_tau_pi.get())
        etiqueta_tau_pi.config(text ="t_pi = "+ "%.3f" %tau_pi)
    except ValueError:
        etiqueta_tau_pi.config(text ="Invalid value")
        
def click_cita_1():
    try:
        cita_1 = float(entrada_cita_1.get())
        etiqueta_cita_1.config(text="c_1 = "+"%.3f" %cita_1)
    except ValueError:
        etiqueta_cita_1.config(text="Invalid value")
        
def click_cita_2():
    try:
        cita_2 = float(entrada_cita_2.get())
        etiqueta_cita_2.config(text="c_2 = "+"%.3f" %cita_2)
    except ValueError:
        etiqueta_cita_2.config(text="Invalid value")
        
def click_cita_3():
    try:
        cita_3 = float(entrada_cita_3.get())
        etiqueta_cita_3.config(text="c_3 = "+"%.3f" %cita_3)
    except ValueError:
        etiqueta_cita_3.config(text="Invalid value")
def click_cita_4():
    try:
        cita_4 = float(entrada_cita_4.get())
        etiqueta_cita_4.config(text="c_4 = " +"%.3f" %cita_4)
    except ValueError:
        etiqueta_cita_4.config(text="Invalid value")

#Buttons Preferences
def click_gamma():
    try:
        gamma = float(entrada_gamma.get())
        etiqueta_gamma.config(text="gamma = "+"%.3f" %gamma)
    except ValueError:
        etiqueta_gamma.config(text="Invalid value")

def click_rho():
    try:
        rho = float(entrada_rho.get())
        etiqueta_rho.config(text="rho = "+"%.3f" %rho)
    except ValueError:
        etiqueta_rho.config(text="Invalid value")
        
def click_beta():
    try:
        beta = float(entrada_beta.get())
        etiqueta_beta.config(text="beta = "+"%.3f" %beta)
    except ValueError:
        etiqueta_beta.config(text="Invalid value")

def click_omega():
    try:
        omega = float(entrada_omega.get())
        etiqueta_omega.config(text="omega = "+"%.3f" %omega)
    except ValueError:
        etiqueta_omega.config(text="Invalid value")
        
#Buttons Technology
def click_A():
    try:
        A = float(entrada_A.get())
        etiqueta_A.config(text="A = "+ "%.3f" %A)
    except ValueError:
        etiqueta_A.config(text="Invalid value")

def click_alpha_p():
    try:
        alpha_p = float(entrada_alpha_p.get())
        etiqueta_alpha_p.config(text="a_p = "+"%.3f" %alpha_p)
    except ValueError:
        etiqueta_alpha_p.config(text="Invalid value")

def click_alpha_g():
    try:
        alpha_g = float(entrada_alpha_g.get())
        etiqueta_alpha_g.config(text="a_g = "+"%.3f" %alpha_g)
    except ValueError:
        etiqueta_alpha_g.config(text="Invalid value")

def click_alpha_l():
    try:
        alpha_l = float(entrada_alpha_l.get())
        etiqueta_alpha_l.config(text="a_l = "+"%.3f" %alpha_l)
    except ValueError:
        etiqueta_alpha_l.config(text="Invalid value")

def click_eta():
    try:
        eta = float(entrada_eta.get())
        etiqueta_eta.config(text="eta = "+"%.3f" %eta)
    except ValueError:
        etiqueta_eta.config(text="Invalid value")

def click_mu():
    try:
        mu = float(entrada_mu.get())
        etiqueta_mu.config(text="mu = "+"%.3f" %mu)
    except ValueError:
        etiqueta_mu.config(text="Invalid value")

def click_delta_p():
    try:
        delta_p = float(entrada_delta_p.get())
        etiqueta_delta_p.config(text="d_p = "+"%.3f" %delta_p)
    except ValueError:
        etiqueta_delta_p.config(text="Invalid value")

def click_delta_g():
    try:
        delta_g = float(entrada_delta_g.get())
        etiqueta_delta_g.config(text="d_g = "+"%.3f" %delta_g)
    except ValueError:
        etiqueta_delta_g.config(text="Invalid value")
        
#Buttons Environment        
def click_debt():
    try:
        debt = float(entrada_debt.get())
        etiqueta_debt.config(text="B = "+"%.3f" %debt)
    except ValueError:
        etiqueta_debt.config(text="Invalid value")

def click_bund_r():
    try:
        bund_r = float(entrada_bund_r.get())
        etiqueta_bund_r.config(text="bund_r = "+"%.3f" %bund_r)
    except ValueError:
        etiqueta_bund_r.config(text="Invalid value")

def click_risk_p():
    try:
        risk_p = float(entrada_risk_p.get())
        etiqueta_risk_p.config(text="r_p = "+"%.3f" %risk_p)
    except ValueError:
        etiqueta_risk_p.config(text="Invalid value")

def check_values():
    count = 0
    alpha_p = float(entrada_alpha_p.get())
    alpha_g = float(entrada_alpha_g.get())
    alpha_l = float(entrada_alpha_l.get())
    if (alpha_p+alpha_g+alpha_l == 1.0):
        count = count +1
        pass
    else:
        tkMessageBox.showinfo('Sum error','alphas must add up to one')

    cita_1 = float(entrada_cita_1.get())
    cita_2 = float(entrada_cita_2.get())
    cita_3 = float(entrada_cita_3.get())
    cita_4 = float(entrada_cita_4.get())
    if (cita_1+cita_2+cita_3+cita_4 == 1.0):
        count = count +1
        pass
    else:
        tkMessageBox.showinfo('Sum error','citas must add up to one')
        
    rho = float(entrada_rho.get())
    if (rho <= -1.0):
        count = count +1
        pass
    else:
        tkMessageBox.showinfo('rho error','rho < -1 required')

    beta = float(entrada_beta.get())
    if (beta < 1.0 and beta > 0):
        count = count +1
        pass
    else:
        tkMessageBox.showinfo('beta error','beta in (0,1) required')

    A = float(entrada_A.get())
    if (A > 0):
        count = count +1
        pass
    else:
        tkMessageBox.showinfo('A error','A > 0 required')
            
    mu = float(entrada_mu.get())
    if (mu < 1.0 and mu > 0):
        count = count +1
        pass
    else:
        tkMessageBox.showinfo('mu error','mu in (0,1) required')

    eta = float(entrada_eta.get())
    if (eta > 0):
        count = count +1
        pass
    else:
        tkMessageBox.showinfo('eta error','eta > 0 required')

    delta_p = float(entrada_delta_p.get())
    if (delta_p > 0):
        count = count +1
        pass
    else:
        tkMessageBox.showinfo('delta_p error','delta_p > 0 required')
            
    delta_g = float(entrada_delta_g.get())
    if (delta_g > 0):
        count = count +1
        pass
    else:
        tkMessageBox.showinfo('delta_g error','delta_g > 0 required')

    alpha_p = float(entrada_alpha_p.get())
    if (alpha_p > 0 and alpha_p < 1):
        count = count +1
        pass
    else:
        tkMessageBox.showinfo('alpha_p error','alpha_p in (0,1) required')

    alpha_g = float(entrada_alpha_g.get())
    if (alpha_g > 0 and alpha_g < 1):
        count = count +1
        pass
    else:
        tkMessageBox.showinfo('alpha_g error','alpha_g in (0,1) required')
        
    alpha_l = float(entrada_alpha_l.get())
    if (alpha_l > 0 and alpha_l < 1):
        count = count +1
        pass
    else:
        tkMessageBox.showinfo('alpha_l error','alpha_l in (0,1) required')

    cita_1 = float(entrada_cita_1.get())
    if (cita_1 > 0 and cita_1 < 1):
        count = count +1
        pass
    else:
        tkMessageBox.showinfo('cita_1 error','cita_1 in (0,1) required')

    cita_2 = float(entrada_cita_2.get())
    if (cita_2 > 0 and cita_2 < 1):
        count = count +1
        pass
    else:
        tkMessageBox.showinfo('cita_2 error','cita_2 in (0,1) required')

    cita_3 = float(entrada_cita_3.get())
    if (cita_3 > 0 and cita_3 < 1):
        count = count +1
        pass
    else:
        tkMessageBox.showinfo('cita_3 error','cita_3 in (0,1) required')

    cita_4 = float(entrada_cita_4.get())
    if (cita_4 > 0 and cita_4 < 1):
        count = count +1
        pass
    else:
        tkMessageBox.showinfo('cita_4 error','cita_4 in (0,1) required')
        
        
    if count == 16:
        #os.system("./GreekCalibrator.py")
        run_program()
    else:
        print count
        pass
    

def call_back_gov(sheet_name):
    sheet = wb.get_sheet_by_name(sheet_name)
    entrada_tau_k.delete(0, END)
    entrada_tau_k.insert(END, sheet['B2'].value)
    entrada_tau_l.delete(0, END)
    entrada_tau_l.insert(END, sheet['B3'].value)
    entrada_tau_c.delete(0, END)
    entrada_tau_c.insert(END, sheet['B4'].value)
    entrada_tau_ss.delete(0, END)
    entrada_tau_ss.insert(END, sheet['B5'].value)
    entrada_tau_pi.delete(0, END)
    entrada_tau_pi.insert(END, sheet['B6'].value)
    entrada_cita_1.delete(0, END)
    entrada_cita_1.insert(END, sheet['B7'].value)
    entrada_cita_2.delete(0, END)
    entrada_cita_2.insert(END, sheet['B8'].value)
    entrada_cita_3.delete(0, END)
    entrada_cita_3.insert(END, sheet['B9'].value)
    entrada_cita_4.delete(0, END)
    entrada_cita_4.insert(END, sheet['B10'].value)

def call_back_pref(sheet_name):
    sheet = wb.get_sheet_by_name(sheet_name)
    entrada_gamma.delete(0, END)
    entrada_gamma.insert(END, sheet['B26'].value)
    entrada_beta.delete(0, END)
    entrada_beta.insert(END, sheet['B25'].value)
    entrada_rho.delete(0, END)
    entrada_rho.insert(END, sheet['B23'].value)
    entrada_omega.delete(0, END)
    entrada_omega.insert(END, sheet['B22'].value)
        
def call_back_tech(sheet_name):
    sheet = wb.get_sheet_by_name(sheet_name)
    entrada_alpha_p.delete(0, END)
    entrada_alpha_p.insert(END, sheet['B13'].value)
    entrada_alpha_g.delete(0, END)
    entrada_alpha_g.insert(END, sheet['B14'].value)
    entrada_alpha_l.delete(0, END)
    entrada_alpha_l.insert(END, sheet['B15'].value)
    entrada_A.delete(0, END)
    entrada_A.insert(END, sheet['B12'].value)
    entrada_mu.delete(0, END)
    entrada_mu.insert(END, sheet['B19'].value)
    entrada_eta.delete(0, END)
    entrada_eta.insert(END, sheet['B18'].value)
    entrada_delta_p.delete(0, END)
    entrada_delta_p.insert(END, sheet['B16'].value)
    entrada_delta_g.delete(0, END)
    entrada_delta_g.insert(END, sheet['B17'].value)

def call_back_env(sheet_name):
    sheet = wb.get_sheet_by_name(sheet_name)
    entrada_debt.delete(0, END)
    entrada_debt.insert(END, sheet['B28'].value)
    entrada_bund_r.delete(0, END)
    entrada_bund_r.insert(END, sheet['B29'].value)
    entrada_risk_p.delete(0, END)
    entrada_risk_p.insert(END, sheet['B30'].value)

def call_back_all(sheet_name):
    sheet = wb.get_sheet_by_name(sheet_name)
    call_back_env(sheet_name)
    call_back_tech(sheet_name)
    call_back_pref(sheet_name)
    call_back_gov(sheet_name)

    
#APP >  Main window
app = Tk()
app.title("MoU Calculator") 
app.geometry("700x730")
app.resizable(0,0)



#VP > Principal canvas
vp = Frame(app,  bd = 3, relief=RAISED)
vp.grid(column=1, row=0, padx=(50,50), pady=(10,10))
vp.columnconfigure(0,weight=1)
vp.rowconfigure(0,weight=1)

vp1 = Frame(app,  bd = 3, relief=RAISED)
vp1.grid(column=1, row=10, padx=(50,50), pady=(10,10))

vp2 = Frame(app,  bd = 3, relief=RAISED)
vp2.grid(column=1, row=20, padx=(50,50), pady=(10,10))

vp3 = Frame(app,  bd = 3, relief=RAISED)
vp3.grid(column=1, row=40, padx=(50,50), pady=(10,10))

vp4 = Frame(app,  bd = 3, relief=RAISED)
vp4.grid(column=1, row=50, padx=(50,50), pady=(10,10))


T1 = Message(app, text = "Government parameters")
T1.config(width=200)
T1.grid(row=0, column=0 )

T2 = Message(app, text = "Preferences parameters")
T2.config(width=200)
T2.grid(row=10, column=0 )

T3 = Message(app, text = "Technology parameters")
T3.config(width=200)
T3.grid(row=20, column=0 )

T4 = Message(app, text = "Environment")
T4.config(width=200)
T4.grid(row=40, column=0 )

#Menus
topMenu = Menu(app)
app.config(menu=topMenu)

fileMenu = Menu(topMenu)
topMenu.add_cascade(label="File", menu=fileMenu)
fileMenu.add_command(label="Save Experiment", command=lambda: call_back_save())

env_menu = Menu(fileMenu)
fileMenu.add_cascade(label="Load Environment", menu=env_menu)
env_menu.add_command(label="Greece", command = lambda: call_back_env('Greece'))
env_menu.add_command(label="Germany", command= lambda: call_back_env('Germany'))
env_menu.add_command(label="Other", command= lambda: call_back_env('Other'))

tech_menu = Menu(fileMenu)
fileMenu.add_cascade(label="Load Technology", menu=tech_menu)
tech_menu.add_command(label="Greece", command = lambda: call_back_tech('Greece'))
tech_menu.add_command(label="Germany", command= lambda: call_back_tech('Germany'))
tech_menu.add_command(label="Other", command= lambda: call_back_tech('Other'))

pref_menu = Menu(fileMenu)
fileMenu.add_cascade(label="Load Preferences", menu=pref_menu)
pref_menu.add_command(label="Greece", command = lambda: call_back_pref('Greece'))
pref_menu.add_command(label="Germany", command= lambda: call_back_pref('Germany'))
pref_menu.add_command(label="Other", command= lambda: call_back_pref('Other'))

gov_menu = Menu(fileMenu)
fileMenu.add_cascade(label="Load Government", menu=gov_menu)
gov_menu.add_command(label="Greece", command = lambda: call_back_gov('Greece'))
gov_menu.add_command(label="Germany", command= lambda: call_back_gov('Germany'))
gov_menu.add_command(label="Other", command= lambda: call_back_gov('Other'))

all_menu = Menu(fileMenu)
fileMenu.add_cascade(label="Load all", menu=all_menu)
all_menu.add_command(label="Greece", command = lambda: call_back_all('Greece'))
all_menu.add_command(label="Germany", command= lambda: call_back_all('Germany'))
all_menu.add_command(label="Other", command= lambda: call_back_all('Other'))

fileMenu.add_separator()
fileMenu.add_command(label="Exit", command=app.destroy)

toolsMenu = Menu(topMenu)
topMenu.add_cascade(label="Tools", menu=toolsMenu)
toolsMenu.add_command(label="Run experiment", command=check_values)

#Frame 1 (Government)
etiqueta_tau_k = Label(vp, text="     Value      ")
etiqueta_tau_k.grid(column=2, row=2, sticky=(W,E))

boton_tau_k = Button(vp, text="tau_k", command=click_tau_k)
boton_tau_k.grid(column=1, row=1)

entrada_tau_k = Entry(vp, width =7)
entrada_tau_k.insert(END, '0.164')
entrada_tau_k.grid(column=2, row=1)


etiqueta_tau_l = Label(vp, text="     Value      ")
etiqueta_tau_l.grid(column=2, row=4, sticky=(W,E))

boton_tau_l = Button(vp, text="tau_l", command=click_tau_l)
boton_tau_l.grid(column=1, row=3)

entrada_tau_l = Entry(vp, width =7)
entrada_tau_l.insert(END, '0.41')
entrada_tau_l.grid(column=2, row=3)


etiqueta_tau_c = Label(vp, text="     Value      ")
etiqueta_tau_c.grid(column=2, row=6, sticky=(W,E))

boton_tau_c = Button(vp, text="tau_c", command=click_tau_c)
boton_tau_c.grid(column=1, row=5)

entrada_tau_c = Entry(vp, width =7)
entrada_tau_c.insert(END, '0.148')
entrada_tau_c.grid(column=2, row=5)


etiqueta_tau_ss = Label(vp, text="     Value      ")
etiqueta_tau_ss.grid(column=2, row=8, sticky=(W,E))

boton_tau_ss = Button(vp, text="tau_ss", command=click_tau_ss)
boton_tau_ss.grid(column=1, row=7)

entrada_tau_ss = Entry(vp, width =7)
entrada_tau_ss.insert(END, '0.356')
entrada_tau_ss.grid(column=2, row=7)


etiqueta_tau_pi = Label(vp, text="     Value      ")
etiqueta_tau_pi.grid(column=2, row=10, sticky=(W,E))

boton_tau_pi = Button(vp, text="tau_pi", command=click_tau_pi)
boton_tau_pi.grid(column=1, row=9)

entrada_tau_pi = Entry(vp, width =7)
entrada_tau_pi.insert(END, '0.25')
entrada_tau_pi.grid(column=2, row=9)


etiqueta_cita_1 = Label(vp, text="     Value      ")
etiqueta_cita_1.grid(column=6, row=2, sticky=(W,E))

boton_cita_1 = Button(vp, text="cita_1", command=click_cita_1)
boton_cita_1.grid(column=5, row=1)

entrada_cita_1 = Entry(vp, width =7)
entrada_cita_1.insert(END, '0.4467')
entrada_cita_1.grid(column=6, row=1)


etiqueta_cita_2 = Label(vp, text="     Value      ")
etiqueta_cita_2.grid(column=6, row=4, sticky=(W,E))

boton_cita_2 = Button(vp, text="cita_2", command=click_cita_2)
boton_cita_2.grid(column=5, row=3)

entrada_cita_2 = Entry(vp, width =7)
entrada_cita_2.insert(END, '0.074')
entrada_cita_2.grid(column=6, row=3)


etiqueta_cita_3 = Label(vp, text="     Value      ")
etiqueta_cita_3.grid(column=6, row=6, sticky=(W,E))

boton_cita_3 = Button(vp, text="cita_3", command=click_cita_3)
boton_cita_3.grid(column=5, row=5)

entrada_cita_3 = Entry(vp, width =7)
entrada_cita_3.insert(END, '0.3246')
entrada_cita_3.grid(column=6, row=5)


etiqueta_cita_4 = Label(vp, text="     Value      ")
etiqueta_cita_4.grid(column=6, row=8, sticky=(W,E))

boton_cita_4 = Button(vp, text='cita_4', command=click_cita_4)
boton_cita_4.grid(column=5, row=7)

entrada_cita_4 = Entry(vp, width = 7)
entrada_cita_4.insert(END, '0.1547')
entrada_cita_4.grid(column=6, row=7)



#Frame 2 (Preferences)
etiqueta_gamma = Label(vp1, text="     Value      ")
etiqueta_gamma.grid(column=2, row=2, sticky=(W,E))

boton_gamma = Button(vp1, text="gamma", command=click_gamma)
boton_gamma.grid(column=1, row=1)

entrada_gamma = Entry(vp1, width =7)
entrada_gamma.insert(END, '0.8361')
entrada_gamma.grid(column=2, row=1)

etiqueta_rho = Label(vp1, text="     Value      ")
etiqueta_rho.grid(column=4, row=2, sticky=(W,E))

boton_rho = Button(vp1, text="rho", command=click_rho)
boton_rho.grid(column=3, row=1)

entrada_rho = Entry(vp1, width =7)
entrada_rho.insert(END, '-1.0')
entrada_rho.grid(column=4, row=1)

etiqueta_beta = Label(vp1, text="     Value      ")
etiqueta_beta.grid(column=2, row=4, sticky=(W,E))

boton_beta = Button(vp1, text="beta", command=click_beta)
boton_beta.grid(column=1, row=3)

entrada_beta = Entry(vp1, width =7)
entrada_beta.insert(END, '0.9606')
entrada_beta.grid(column=2, row=3)

etiqueta_omega = Label(vp1, text="     Value      ")
etiqueta_omega.grid(column=4, row=4, sticky=(W,E))

boton_omega = Button(vp1, text="omega", command=click_omega)
boton_omega.grid(column=3, row=3)

entrada_omega = Entry(vp1, width =7)
entrada_omega.insert(END, '0.0806')
entrada_omega.grid(column=4, row=3)


#Frame 3 (Technology)
etiqueta_alpha_p = Label(vp2, text="     Value      ")
etiqueta_alpha_p.grid(column=2, row=4, sticky=(W,E))

boton_alpha_p = Button(vp2, text="alpha_p", command=click_alpha_p)
boton_alpha_p.grid(column=1, row=3)

entrada_alpha_p = Entry(vp2, width =7)
entrada_alpha_p.insert(END, '0.3065')
entrada_alpha_p.grid(column=2, row=3)


etiqueta_alpha_g = Label(vp2, text="     Value      ")
etiqueta_alpha_g.grid(column=2, row=6, sticky=(W,E))

boton_alpha_g = Button(vp2, text="alpha_g", command=click_alpha_g)
boton_alpha_g.grid(column=1, row=5)

entrada_alpha_g = Entry(vp2, width =7)
entrada_alpha_g.insert(END, '0.1082')
entrada_alpha_g.grid(column=2, row=5)


etiqueta_alpha_l = Label(vp2, text="     Value      ")
etiqueta_alpha_l.grid(column=2, row=8, sticky=(W,E))

boton_alpha_l = Button(vp2, text="alpha_l", command=click_alpha_l)
boton_alpha_l.grid(column=1, row=7)

entrada_alpha_l = Entry(vp2, width =7)
entrada_alpha_l.insert(END, '0.5853')
entrada_alpha_l.grid(column=2, row=7)


etiqueta_eta = Label(vp2, text="     Value      ")
etiqueta_eta.grid(column=6, row=8, sticky=(W,E))

boton_eta = Button(vp2, text="eta", command=click_eta)
boton_eta.grid(column=5, row=7)

entrada_eta = Entry(vp2, width =7)
entrada_eta.insert(END, '0.47789')
entrada_eta.grid(column=6, row=7)


etiqueta_mu = Label(vp2, text="     Value      ")
etiqueta_mu.grid(column=6, row=6, sticky=(W,E))

boton_mu = Button(vp2, text="mu", command=click_mu)
boton_mu.grid(column=5, row=5)

entrada_mu = Entry(vp2, width =7)
entrada_mu.insert(END, '0.6008')
entrada_mu.grid(column=6, row=5)

etiqueta_A = Label(vp2, text="     Value      ")
etiqueta_A.grid(column=6, row=4, sticky=(W,E))

boton_A = Button(vp2, text="A", command=click_A)
boton_A.grid(column=5, row=3)

entrada_A = Entry(vp2, width =7)
entrada_A.insert(END, '1.6044')
entrada_A.grid(column=6, row=3)

etiqueta_delta_p = Label(vp2, text="     Value      ")
etiqueta_delta_p.grid(column=8, row=4, sticky=(W,E))

boton_delta_p = Button(vp2, text="delta_p", command=click_delta_p)
boton_delta_p.grid(column=7, row=3)

entrada_delta_p = Entry(vp2, width =7)
entrada_delta_p.insert(END, '0.08')
entrada_delta_p.grid(column=8, row=3)

etiqueta_delta_g = Label(vp2, text="     Value      ")
etiqueta_delta_g.grid(column=8, row=6, sticky=(W,E))

boton_delta_g = Button(vp2, text="delta_g", command=click_delta_g)
boton_delta_g.grid(column=7, row=5)

entrada_delta_g = Entry(vp2, width =7)
entrada_delta_g.insert(END, '0.04')
entrada_delta_g.grid(column=8, row=5)

#Frame 4 (Environment)

etiqueta_debt = Label(vp3, text="     Value      ")
etiqueta_debt.grid(column=2, row=2, sticky=(W,E))

boton_debt = Button(vp3, text="B", command=click_debt)
boton_debt.grid(column=1, row=1)

entrada_debt = Entry(vp3, width =7)
entrada_debt.insert(END, '110')
entrada_debt.grid(column=2, row=1)

etiqueta_bund_r = Label(vp3, text="     Value      ")
etiqueta_bund_r.grid(column=6, row=2, sticky=(W,E))

boton_bund_r = Button(vp3, text="Bund R", command=click_bund_r)
boton_bund_r.grid(column=5, row=1)

entrada_bund_r = Entry(vp3, width =7)
entrada_bund_r.insert(END, '0.041')
entrada_bund_r.grid(column=6, row=1)

etiqueta_risk_p = Label(vp3, text="     Value      ")
etiqueta_risk_p.grid(column=6, row=4, sticky=(W,E))

boton_risk_p = Button(vp3, text="Premium", command=click_risk_p)
boton_risk_p.grid(column=5, row=3)

entrada_risk_p = Entry(vp3, width =7)
entrada_risk_p.insert(END, '0.0')
entrada_risk_p.grid(column=6, row=3)


def run_program():
    #Basic data and calibration for 2002-2006
    #Env
    bund_r = float(entrada_bund_r.get())
    risk_p = float(entrada_risk_p.get())
    B = float(entrada_debt.get())
    RB = bund_r+risk_p
    RB1 = RB
    #Gov
    tauk = float(entrada_tau_k.get())
    taus = float(entrada_tau_ss.get())
    tauc = float(entrada_tau_c.get())
    taul = float(entrada_tau_l.get())
    taupi = float(entrada_tau_pi.get())
    theta1 = float(entrada_cita_1.get())
    theta2 = float(entrada_cita_2.get())
    theta3 = float(entrada_cita_3.get())
    theta4 = float(entrada_cita_4.get())
    #Pref
    beta = float(entrada_beta.get())
    rho = float(entrada_rho.get())
    omega = float(entrada_omega.get())
    gamma = float(entrada_gamma.get())
    #Tech
    deltap = float(entrada_delta_p.get())
    deltag = float(entrada_delta_g.get())
    alphap = float(entrada_alpha_p.get())
    alphag = float(entrada_alpha_g.get())
    alphal = float(entrada_alpha_l.get())
    eta = float(entrada_eta.get())
    mu = float(entrada_mu.get())
    A = float(entrada_A.get())
    #Misc
    pi_c = 1
    H = 100
    
    #Compute a steady state for each value of G/Y
    tic=timeit.default_timer()
    init = 0.01    #%Lowest G/Y ratio
    final = 0.65    # %Highest G/Y ratio
    T = 1000    #%Density 1/T

    def g(x):
        return GreeceModelRBFOC(x, param)

    varNames = ["Kpss", "Kgss", "Lpss", "Lgss", "Bss", "Ipss", "Igss", "Lss", \
            "Yss", "Rss", "PmKgss", "Wpss", "PmLgss", "Gss", "Cgss", "Zss", \
            "Cpss", "Css", "Wgss", "PIss", "IFss", "ratGY"]

    for name in varNames:
        globals()[name] = np.zeros(T)
    
    x0 = [237, 83, 46, 11, 110]
    for t in range(T):
        ratGY[t] = init+(final-init)*t/(T-1)
        param = [alphap, alphag, RB1, deltap, deltag, gamma, rho, theta1, \
             theta2, theta3, pi_c, mu, eta, omega, tauc, taul, tauk, taus, \
             taupi, H, A, ratGY[t], RB]
    
        crit = 1e-10
        maxit = 1000
        sol = fsolve(g, x0, xtol=1.e-06)
        Kpss[t] = sol[0]
        Kgss[t] = sol[1]
        Lpss[t] = sol[2]
        Lgss[t] = sol[3]
        Bss[t] = sol[4]
        Ipss[t] = deltap*Kpss[t]
        Igss[t] = deltag*Kgss[t]
        Lss[t] = Lgss[t]+Lpss[t]
        Yss[t] = A*Kpss[t]**alphap*Kgss[t]**alphag*(mu*Lpss[t]**eta+ \
            (1-mu)*Lgss[t]**eta)**(alphal/eta)
        Rss[t] = alphap*A*Kpss[t]**(alphap-1)*Kgss[t]**alphag*(mu*Lpss[t]** \
            eta+(1-mu)*Lgss[t]**eta)**(alphal/eta)
        PmKgss[t] = alphag*A*Kpss[t]**alphap*Kgss[t]**(alphag-1)*(mu*Lpss[t]** \
                eta+(1-mu)*Lgss[t]**eta)**(alphal/eta)
        Wpss[t] = (alphal*A*mu*Lpss[t]**(eta-1)*Kpss[t]**alphap*Kgss[t]**alphag* \
            (mu*Lpss[t]**eta+(1-mu)*Lgss[t]**eta)**(alphal/eta-1))/(1+taus)
        PmLgss[t] = (alphal*A*(1-mu)*Lgss[t]**(eta-1)*Kpss[t]**alphap*Kgss[t]** \
            alphag*(mu*Lpss[t]**eta+(1-mu)*Lgss[t]**eta)** \
            (alphal/eta-1))/(1+taus)
        Gss[t] = ratGY[t]*Yss[t]
        Cgss[t] = theta1*Gss[t]
        Zss[t] = theta4*Gss[t]
        Cpss[t] = Yss[t]-Ipss[t]-Igss[t]-Cgss[t]-RB*Bss[t]
        Css[t] = Cpss[t]+pi_c*Cgss[t]
        Wgss[t] = (omega/(1-omega))**(-1/(2*rho))*(theta3*Gss[t]/(1+taus))**(1/2)
        PIss[t] = Yss[t]-(1+taus)*Wpss[t]*Lpss[t]-Rss[t]*Kpss[t]
        IFss[t] = tauc*Cpss[t]+(taul+taus)*(Wpss[t]*Lpss[t]+Wgss[t]*Lgss[t])+ \
              tauk*(Rss[t]-deltap)*Kpss[t]+taupi*PIss[t]
        tb = t
        if Bss[t] <= 0:
            break
        x0 = [Kpss[t], Kgss[t], Lpss[t], Lgss[t], Bss[t]]

    
    def call_back_save():
        wb = Workbook()
        dest_filename = 'Experiment_Results.xlsx'
        ws1 = wb.create_sheet(title="Experiment")
        #ws1 = wb.active
        
        
        for row in range(2,tb):
            _= ws1.cell(column = 1, row = 1, value = 'Kpss')
            _= ws1.cell(column = 1, row = row, value = Kpss[row-2])

            _= ws1.cell(column = 2, row = 1, value = 'Kgss')
            _= ws1.cell(column = 2, row = row, value = Kgss[row-2])

            _= ws1.cell(column = 3, row = 1, value = 'Lpss')            
            _= ws1.cell(column = 3, row = row, value = Lpss[row-2])

            _= ws1.cell(column = 4, row = 1, value = 'Lgss')
            _= ws1.cell(column = 4, row = row, value = Lgss[row-2])

            _= ws1.cell(column = 5, row = 1, value = 'Bss')
            _= ws1.cell(column = 5, row = row, value = Bss[row-2])

            _= ws1.cell(column = 6, row = 1, value = 'Ipss')
            _= ws1.cell(column = 6, row = row, value = Ipss[row-2])

            _= ws1.cell(column = 7, row = 1, value = 'Igss')
            _= ws1.cell(column = 7, row = row, value = Igss[row-2])

            _= ws1.cell(column = 8, row = 1, value = 'Lss')
            _= ws1.cell(column = 8, row = row, value = Lss[row-2])

            _= ws1.cell(column = 9, row = 1, value = 'Yss')
            _= ws1.cell(column = 9, row = row, value = Yss[row-2])

            _= ws1.cell(column = 10, row = 1, value = 'Rss')
            _= ws1.cell(column = 10, row = row, value = Rss[row-2])

            _= ws1.cell(column = 11, row = 1, value = 'Wpss')
            _= ws1.cell(column = 11, row = row, value = Wpss[row-2])

            _= ws1.cell(column = 12, row = 1, value = 'Wgss')
            _= ws1.cell(column = 12, row = row, value = Wgss[row-2])

            _= ws1.cell(column = 13, row = 1, value = 'Gss')
            _= ws1.cell(column = 13, row = row, value = Gss[row-2])

            _= ws1.cell(column = 14, row = 1, value = 'Zss')
            _= ws1.cell(column = 14, row = row, value = Zss[row-2])

            _= ws1.cell(column = 15, row = 1, value = 'Cgss')
            _= ws1.cell(column = 15, row = row, value = Cgss[row-2])

            _= ws1.cell(column = 16, row = 1, value = 'Cpss')
            _= ws1.cell(column = 16, row = row, value = Cpss[row-2])

            _= ws1.cell(column = 17, row = 1, value = 'Css')
            _= ws1.cell(column = 17, row = row, value = Css[row-2])

            _= ws1.cell(column = 18, row = 1, value = 'Zss')
            _= ws1.cell(column = 18, row = row, value = Zss[row-2])

            _= ws1.cell(column = 19, row = 1, value = 'PIss')
            _= ws1.cell(column = 19, row = row, value = PIss[row-2])

            _= ws1.cell(column = 20, row = 1, value = 'IFss')
            _= ws1.cell(column = 20, row = row, value = IFss[row-2])

            _= ws1.cell(column = 21, row = 1, value = 'ratGY')
            _= ws1.cell(column = 21, row = row, value = ratGY[row-2])
            
        wb.save(filename = dest_filename)


    botonSave = Button(vp4, text="Save", command =  call_back_save())
    botonSave.grid(column=6, row=5)

    

    toc=timeit.default_timer()
    textTime = "%.4f" %(toc - tic )
    #time = (toc - tic )
    app.title("MoU computed in " + textTime + " seconds")

    dataGreece = np.array([[2002, 45.1, 101.7],
                       [2003, 44.7, 97.4],
                       [2004, 45.5, 98.6],
                       [2005, 44.6, 100],
                       [2006, 45.3, 106.1],
                       [2007, 47.5, 105.4],
                       [2008, 50.6, 110.7],
                       [2009, 54, 129.7],
                       [2010, 51.5, 148.3],
                       [2011, 51.8, 170.6]])

    fig = plt.figure(num = 1, figsize=(6, 6), dpi = 80, facecolor = '#888888', \
        edgecolor= 'k')
    plt.xlim(40,55)
    plt.ylim(35,180)
    plt.plot(100*Gss[0:tb]/Yss[0:tb], 100*Bss[0:tb]/Yss[0:tb])
    plt.title('Debt and Expenditures')
    plt.plot(dataGreece[:,1], dataGreece[:,2], '-o')
    plt.plot(dataGreece[:,1], dataGreece[:,2], 'ro')
    plt.grid(True)
    plt.text(44, 90, '2002-2006')
    plt.text(47.5, 100, '2007')
    plt.text(51, 108, '2008')
    plt.text(52, 130, '2009')
    plt.text(50, 150, '2010')
    plt.text(50.5, 172, '2011')

    plt.ylabel('Debt to GDP')
    plt.xlabel('Expenditures to GDP')
    canvas = FigureCanvasTkAgg(fig, master=app)
    plot_widget = canvas.get_tk_widget()

    plt.show()
    
#Ejecucion de la app
app.mainloop()
    
